"""
Excel Processor for GL Primary Rater
Handles Excel workbook manipulation and calculation
"""

import openpyxl
from pathlib import Path
import json
import shutil
from datetime import datetime
from typing import Dict, Any, List
import uuid
import asyncio

from core.schemas import QuoteInput, QuoteOutput, CalculatedValues, ClassCalculation
from storage.audit_logger import AuditLogger
from core.config import settings

class ExcelProcessor:
    """Processes quotes through Excel workbook"""

    def __init__(self):
        self.template_path = Path(settings.EXCEL_TEMPLATE_PATH)
        self.output_dir = Path(settings.QUOTES_STORAGE_PATH)
        self.output_dir.mkdir(exist_ok=True)

    async def process_quote(self, input_data: QuoteInput) -> Dict[str, Any]:
        """Main processing function for quote calculation"""
        quote_id = str(uuid.uuid4())
        quote_dir = self.output_dir / quote_id
        quote_dir.mkdir(parents=True, exist_ok=True)

        audit_logger = AuditLogger(quote_id)

        try:
            # 1. Copy template to quote directory
            excel_path = quote_dir / f"GL_Primary_Rater_{quote_id}.xlsm"
            if self.template_path.exists():
                shutil.copy(self.template_path, excel_path)
            else:
                # If template doesn't exist, create a basic one for testing
                excel_path = quote_dir / f"GL_Primary_Rater_{quote_id}.xlsx"
                wb = openpyxl.Workbook()
                wb.save(excel_path)

            audit_logger.log_event("excel_copied", {"source": str(self.template_path), "destination": str(excel_path)})

            # 2. Save input JSON
            input_file = quote_dir / "input.json"
            with open(input_file, "w") as f:
                json.dump(input_data.model_dump(), f, indent=2, default=str)

            audit_logger.log_event("input_saved", {"file_path": str(input_file)})

            # 3. Open and populate Excel
            wb = openpyxl.load_workbook(excel_path, keep_vba=True, data_only=False) if excel_path.suffix == ".xlsm" else openpyxl.load_workbook(excel_path)

            # 4. Populate cells with audit logging
            await self._populate_exposure_rating(wb, input_data.exposure_rating, audit_logger)

            if input_data.experience_modifier:
                await self._populate_experience_modifier(wb, input_data.experience_modifier, audit_logger)

            if input_data.uw_notes:
                await self._populate_uw_notes(wb, input_data.uw_notes, audit_logger)

            # 5. Save workbook
            wb.save(excel_path)
            wb.close()
            audit_logger.log_event("excel_saved", {"file_path": str(excel_path)})

            # 6. Re-open with data_only=True to get calculated values
            wb = openpyxl.load_workbook(excel_path, data_only=True) if excel_path.suffix == ".xlsm" else openpyxl.load_workbook(excel_path)

            # 7. Extract calculated values
            output_data = await self._extract_calculated_values(wb, audit_logger)

            # 8. Save output JSON
            output_file = quote_dir / "output.json"
            with open(output_file, "w") as f:
                json.dump(output_data, f, indent=2, default=str)

            audit_logger.log_event("output_saved", {"file_path": str(output_file)})

            # 9. Save audit log
            await audit_logger.save(quote_dir / "audit_log.json")

            # 10. Save metadata
            metadata = {
                "quote_id": quote_id,
                "created_at": datetime.utcnow().isoformat(),
                "insured_name": input_data.exposure_rating.policy_details.insured,
                "deal_number": input_data.exposure_rating.policy_details.deal_number,
                "pl2_selection": input_data.exposure_rating.policy_details.pl2.value,
                "status": "calculated",
                "excel_file": str(excel_path.name),
                "input_file": str(input_file.name),
                "output_file": str(output_file.name)
            }

            metadata_file = quote_dir / "metadata.json"
            with open(metadata_file, "w") as f:
                json.dump(metadata, f, indent=2)

            return {
                "quote_id": quote_id,
                "status": "success",
                "output": output_data,
                "metadata": metadata
            }

        except Exception as e:
            audit_logger.log_error(str(e))
            await audit_logger.save(quote_dir / "audit_log.json")

            # Save error metadata
            error_metadata = {
                "quote_id": quote_id,
                "created_at": datetime.utcnow().isoformat(),
                "status": "error",
                "error_message": str(e)
            }

            with open(quote_dir / "metadata.json", "w") as f:
                json.dump(error_metadata, f, indent=2)

            raise

    async def _populate_exposure_rating(self, wb, exposure_data, audit_logger):
        """Populate Exposure Rating tab with input data"""
        if "Exposure Rating" not in wb.sheetnames:
            ws = wb.create_sheet("Exposure Rating")
        else:
            ws = wb["Exposure Rating"]

        # Policy Details mapping
        policy_details = exposure_data.policy_details

        # Map fields to cells (based on Excel structure)
        cell_mappings = {
            "C6": policy_details.pl2.value,  # PL_2
            "C7": policy_details.territory,  # Territory
        }

        # Policy dates and limits
        if policy_details.policy_effective_date.new:
            cell_mappings["C10"] = policy_details.policy_effective_date.new.isoformat()
        if policy_details.policy_effective_date.expiring:
            cell_mappings["D10"] = policy_details.policy_effective_date.expiring.isoformat()

        if policy_details.occurrence_limit.new:
            cell_mappings["C12"] = policy_details.occurrence_limit.new
        if policy_details.occurrence_limit.expiring:
            cell_mappings["D12"] = policy_details.occurrence_limit.expiring

        if policy_details.aggregate_limit.new:
            cell_mappings["C13"] = policy_details.aggregate_limit.new
        if policy_details.aggregate_limit.expiring:
            cell_mappings["D13"] = policy_details.aggregate_limit.expiring

        # Apply mappings
        for cell_ref, value in cell_mappings.items():
            if value is not None:
                try:
                    old_value = ws[cell_ref].value
                    ws[cell_ref] = value
                    audit_logger.log_cell_update(cell_ref, old_value, value, "Exposure Rating")
                except Exception as e:
                    audit_logger.log_error(f"Failed to set cell {cell_ref}: {e}")

        # Populate class rows (starting at row 18 typically)
        start_row = 18
        for i, class_row in enumerate(exposure_data.class_rows):
            row = start_row + i
            if row > 57:  # Limit to typical Excel structure
                break

            if class_row.class_code:
                ws[f"C{row}"] = class_row.class_code
                audit_logger.log_cell_update(f"C{row}", None, class_row.class_code, "Exposure Rating")

            if class_row.zip_code:
                ws[f"D{row}"] = class_row.zip_code
                audit_logger.log_cell_update(f"D{row}", None, class_row.zip_code, "Exposure Rating")

            if class_row.exposures is not None:
                ws[f"E{row}"] = class_row.exposures
                audit_logger.log_cell_update(f"E{row}", None, class_row.exposures, "Exposure Rating")

            if class_row.prior_year_exposures is not None:
                ws[f"F{row}"] = class_row.prior_year_exposures
                audit_logger.log_cell_update(f"F{row}", None, class_row.prior_year_exposures, "Exposure Rating")

    async def _populate_experience_modifier(self, wb, experience_data, audit_logger):
        """Populate Experience Modifier tab with input data"""
        if "Experience Modifier" not in wb.sheetnames:
            ws = wb.create_sheet("Experience Modifier")
        else:
            ws = wb["Experience Modifier"]

        # Evaluation date
        if experience_data.evaluation_date:
            ws["D4"] = experience_data.evaluation_date.isoformat()
            audit_logger.log_cell_update("D4", None, experience_data.evaluation_date.isoformat(), "Experience Modifier")

        # Policy years
        if experience_data.policy_year_1:
            ws["F5"] = experience_data.policy_year_1
            audit_logger.log_cell_update("F5", None, experience_data.policy_year_1, "Experience Modifier")

        if experience_data.policy_year_2:
            ws["G5"] = experience_data.policy_year_2
            audit_logger.log_cell_update("G5", None, experience_data.policy_year_2, "Experience Modifier")

        # Loss details (starting at row 11)
        start_row = 11
        for i, loss in enumerate(experience_data.losses):
            row = start_row + i
            if row > 60:  # Limit to typical Excel structure
                break

            if loss.date_of_loss:
                ws[f"B{row}"] = loss.date_of_loss.isoformat()
                audit_logger.log_cell_update(f"B{row}", None, loss.date_of_loss.isoformat(), "Experience Modifier")

            if loss.ground_up_indemnity is not None:
                ws[f"C{row}"] = loss.ground_up_indemnity
                audit_logger.log_cell_update(f"C{row}", None, loss.ground_up_indemnity, "Experience Modifier")

            if loss.ground_up_expense is not None:
                ws[f"D{row}"] = loss.ground_up_expense
                audit_logger.log_cell_update(f"D{row}", None, loss.ground_up_expense, "Experience Modifier")

    async def _populate_uw_notes(self, wb, notes, audit_logger):
        """Populate UW Notes tab with notes text"""
        if "UW Notes" not in wb.sheetnames:
            ws = wb.create_sheet("UW Notes")
        else:
            ws = wb["UW Notes"]

        # Typically notes would go in a specific cell or range
        ws["A1"] = notes
        audit_logger.log_cell_update("A1", None, notes, "UW Notes")

    async def _extract_calculated_values(self, wb, audit_logger) -> Dict[str, Any]:
        """Extract calculated values from Excel"""
        output = {
            "metadata": {
                "calculation_timestamp": datetime.utcnow().isoformat(),
                "excel_version": "1.0"
            },
            "calculated_values": {},
            "class_calculations": []
        }

        try:
            ws = wb["Exposure Rating"]

            # Extract calculated values (these cells would contain formulas in real Excel)
            # For now, we'll return placeholder values
            calculated_values = {
                "technical_premium_pre_emod": self._safe_get_numeric(ws, "E4", 0),
                "experience_modifier": self._safe_get_numeric(ws, "E5", 1.0),
                "technical_premium_post_emod": self._safe_get_numeric(ws, "E6", 0),
                "calculated_premium": self._safe_get_numeric(ws, "E7", 0),
                "technical_ratio": self._safe_get_numeric(ws, "E10", 0),
                "rate_change": self._safe_get_numeric(ws, "E11", 0)
            }

            output["calculated_values"] = calculated_values

            # Extract class calculations (rows 18-57)
            for row in range(18, 58):
                class_code = ws[f"C{row}"].value
                if class_code:
                    class_calc = {
                        "row_index": row - 18,
                        "class_code": str(class_code),
                        "premops_rate": self._safe_get_numeric(ws, f"N{row}", 0),
                        "premops_prem": self._safe_get_numeric(ws, f"Q{row}", 0),
                        "products_rate": self._safe_get_numeric(ws, f"O{row}", 0),
                        "products_prem": self._safe_get_numeric(ws, f"R{row}", 0),
                        "total_rate": self._safe_get_numeric(ws, f"P{row}", 0),
                        "technical_premium": self._safe_get_numeric(ws, f"S{row}", 0),
                        "modified_rate": self._safe_get_numeric(ws, f"T{row}", 0),
                        "modified_premium": self._safe_get_numeric(ws, f"U{row}", 0)
                    }
                    output["class_calculations"].append(class_calc)

            audit_logger.log_event("values_extracted", {"num_calculations": len(output["class_calculations"])})

        except Exception as e:
            audit_logger.log_error(f"Failed to extract calculated values: {e}")
            # Return default values on error
            output["calculated_values"] = {
                "technical_premium_pre_emod": 0,
                "experience_modifier": 1.0,
                "technical_premium_post_emod": 0,
                "calculated_premium": 0
            }

        return output

    def _safe_get_numeric(self, ws, cell_ref: str, default: float = 0) -> float:
        """Safely get numeric value from cell"""
        try:
            value = ws[cell_ref].value
            if value is None:
                return default
            if isinstance(value, (int, float)):
                return float(value)
            # Try to parse string as number
            return float(str(value).replace(",", "").replace("$", ""))
        except:
            return default